"""5-sheet audit workbook builder.

Sheet 1 is a marketing-quality cover (flyer aesthetic — hidden gridlines,
embedded logo, brand-blue panels, native bar chart, set up to print as a
single US-Letter portrait page). Sheets 2-5 follow the permit-audit skill
formatting spec verbatim.
"""
from __future__ import annotations

import io
import os
from datetime import date, datetime
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice, PatternFillProperties
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

from .cross_reference import (
    AuditData,
    STATUS_PARKING_ONLY,
    STATUS_REGISTERED,
    STATUS_TOW_RISK,
    STATUS_VACANT_HAS_PERMITS,
    STATUS_VACANT_OK,
)

ASSETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
APOLLO_LOGO_PATH = os.path.join(ASSETS_DIR, "apollo-logo.png")


# ------------------------------------------------------------------
# Color palette — straight from SKILL.md
# ------------------------------------------------------------------
BRAND_BLUE = "3F66A0"        # cover-sheet brand color (per spec §3a)
BRAND_BLUE_DARK = "2F4E7A"
DARK_BLUE = "2F5496"
MEDIUM_BLUE = "4472C4"
GREEN_FILL = "C6EFCE"
GREEN_TEXT = "006100"
RED_FILL = "FFC7CE"
RED_TEXT = "9C0006"
YELLOW_FILL = "FFEB9C"
YELLOW_TEXT = "9C6500"
ORANGE_HIGHLIGHT = "FFF2CC"
ALT_ROW = "F2F2F2"
NAVY = "1F4E79"
BORDER_GRAY = "D9D9D9"
RED_BRIGHT = "FF0000"
RED_DEEP = "C00000"
SUBTITLE_GRAY = "404040"
GENERATED_GRAY = "808080"

# ------------------------------------------------------------------
# Reusable style fragments
# ------------------------------------------------------------------
ARIAL = "Arial"
THIN = Side(style="thin", color=BORDER_GRAY)
DATA_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITE = "FFFFFF"


def _header_font() -> Font:
    return Font(name=ARIAL, size=11, bold=True, color=WHITE)


def _data_font() -> Font:
    return Font(name=ARIAL, size=10)


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _set_columns(ws: Worksheet, widths: Iterable[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _apply_header_row(ws: Worksheet, headers: Iterable[str], fill_color: str) -> None:
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.fill = _fill(fill_color)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = DATA_BORDER


def _apply_alt_shading(ws: Worksheet, start_row: int, end_row: int, ncols: int) -> None:
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = _fill(ALT_ROW)


def _set_data_borders(ws: Worksheet, start_row: int, end_row: int, ncols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = DATA_BORDER
            if not cell.font or cell.font.name is None:
                cell.font = _data_font()


# ------------------------------------------------------------------
# Public entry point
# ------------------------------------------------------------------
def build_workbook(audit: AuditData) -> bytes:
    wb = Workbook()
    # Drop the default sheet — we make our own
    default = wb.active
    wb.remove(default)

    _build_cover(wb, audit)
    _build_cross_reference(wb, audit)
    _build_vehicle_details(wb, audit)
    _build_vacant_future(wb, audit)
    _build_discrepancies(wb, audit)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------
# Sheet 1 — Cover (marketing-quality, single-page US-Letter portrait)
# Spec §3a: hide gridlines + headers, embed logo, brand-blue panels,
# native bar chart, generous whitespace. Treat as a flyer.
# ------------------------------------------------------------------
COVER_TITLE = "PARKING AUDIT"


def _cover_fmt_date(iso_or_none: Optional[str]) -> str:
    """Format an ISO yyyy-mm-dd date as 'April 26, 2026'. Falls back to today."""
    s = (iso_or_none or "").strip()
    if s:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%B %-d, %Y")
        except ValueError:
            pass
    return date.today().strftime("%B %-d, %Y")


def _band_fill(ws: Worksheet, row: int, start_col: int, end_col: int, *, label: str) -> None:
    """Render a brand-blue section header band spanning [start_col..end_col] on `row`."""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=label)
    cell.fill = _fill(BRAND_BLUE)
    cell.font = Font(name=ARIAL, size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26


def _build_cover(wb: Workbook, audit: AuditData) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_properties.tabColor = BRAND_BLUE

    # Flyer aesthetic — hide gridlines and row/column headers everywhere.
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Page setup — US Letter, portrait, fit to one page.
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)
    ws.print_options.gridLines = False
    ws.print_options.headings = False

    # 8-column grid — wider center columns for property info, narrower edges.
    _set_columns(ws, [3, 12, 14, 14, 14, 14, 12, 3])

    # ── 1) Header band ───────────────────────────────────────────────────
    # Logo top-left. Embedding the PNG; openpyxl renders it via drawingML.
    if os.path.exists(APOLLO_LOGO_PATH):
        try:
            img = XLImage(APOLLO_LOGO_PATH)
            # Source is 5401x1795 (banner). Target ~3.0" wide × 1.0" tall.
            img.width = 220
            img.height = 73
            img.anchor = "B2"
            ws.add_image(img)
        except Exception:  # noqa: BLE001 — logo is decorative; never fail audit on it
            pass
    ws.row_dimensions[2].height = 70
    # Document title — large bold uppercase, right side of header
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7)
    title = ws.cell(row=2, column=4, value=COVER_TITLE)
    title.font = Font(name=ARIAL, size=28, bold=True, color=BRAND_BLUE)
    title.alignment = Alignment(horizontal="right", vertical="center")

    # Brand-blue accent rule beneath the header
    ws.row_dimensions[3].height = 5
    for c in range(2, 8):
        ws.cell(row=3, column=c).fill = _fill(BRAND_BLUE)

    # ── 2) Identification block ─────────────────────────────────────────
    ws.row_dimensions[5].height = 14
    label_prepared = ws.cell(row=5, column=2, value="Prepared for:")
    label_prepared.font = Font(name=ARIAL, size=10, color=SUBTITLE_GRAY)
    label_prepared.alignment = Alignment(horizontal="left")

    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=7)
    prop = ws.cell(row=6, column=2, value=audit.property_name)
    prop.font = Font(name=ARIAL, size=20, bold=True, color=BRAND_BLUE)
    prop.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[6].height = 28

    if audit.property_address:
        ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=7)
        addr = ws.cell(row=7, column=2, value=audit.property_address)
        addr.font = Font(name=ARIAL, size=11, color="595959")
        addr.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[7].height = 18

    audit_date_str = _cover_fmt_date(audit.audit_date)
    audit_date_row = 8 if audit.property_address else 7
    ws.merge_cells(
        start_row=audit_date_row, start_column=2,
        end_row=audit_date_row, end_column=7,
    )
    d = ws.cell(row=audit_date_row, column=2, value=f"Audit Date: {audit_date_str}")
    d.font = Font(name=ARIAL, size=10, color=SUBTITLE_GRAY)
    d.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[audit_date_row].height = 16

    # Spacer row before Property Information panel
    spacer_row = audit_date_row + 1
    ws.row_dimensions[spacer_row].height = 14

    # ── 3) Property Information panel ───────────────────────────────────
    pi_band = spacer_row + 1
    _band_fill(ws, pi_band, 2, 7, label="PROPERTY INFORMATION")

    pl = audit.parking_lot or {}
    specialty_tags_str = (
        f"Total Specialty Spaces ({', '.join(audit.specialty_tags)})"
        if audit.specialty_tags else "Total Specialty Spaces"
    )

    # Two-column data grid:
    # Left column (rows): Number of Units, Reserved/Carport, Open, Guest, Handicap, Specialty
    # Right column (rows): Occupied Units, Total Parking Spaces
    left_rows = [
        ("Number of Units", audit.num_units_form or audit.rent_roll_units),
        ("Total Reserved/Carport Spaces", pl.get("reserved", 0)),
        ("Total Open Parking Spaces", pl.get("open", 0)),
        ("Total Guest Spaces", pl.get("guest", 0)),
        ("Total Handicap Spaces", pl.get("handicap", 0)),
        (specialty_tags_str, pl.get("specialty", 0)),
    ]
    right_rows = [
        ("Occupied Units", audit.occupied),
        ("Total Parking Spaces", pl.get("total", 0)),
    ]

    pi_start = pi_band + 1
    for i, (label, value) in enumerate(left_rows):
        r = pi_start + i
        ws.row_dimensions[r].height = 22
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        lbl = ws.cell(row=r, column=2, value=label)
        lbl.font = Font(name=ARIAL, size=10, color="404040")
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val = ws.cell(row=r, column=4, value=value)
        val.font = Font(name=ARIAL, size=11, bold=True, color="0F172A")
        val.alignment = Alignment(horizontal="right", vertical="center")
    for i, (label, value) in enumerate(right_rows):
        r = pi_start + i
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        lbl = ws.cell(row=r, column=5, value=label)
        lbl.font = Font(name=ARIAL, size=10, color="404040")
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val = ws.cell(row=r, column=7, value=value)
        val.font = Font(name=ARIAL, size=11, bold=True, color="0F172A")
        val.alignment = Alignment(horizontal="right", vertical="center")

    pi_end = pi_start + len(left_rows) - 1

    # Spacer
    af_spacer = pi_end + 1
    ws.row_dimensions[af_spacer].height = 14

    # ── 4) Audit Findings panel ─────────────────────────────────────────
    af_band = af_spacer + 1
    _band_fill(ws, af_band, 2, 7, label="AUDIT FINDINGS")

    cars_registered = audit.total_vehicles
    total_spaces = pl.get("total", 0)
    reserved_spaces = pl.get("reserved", 0)
    is_myvip = audit.parking_program == "myvip"
    units_without_myvip = (
        max(audit.occupied - audit.myvip_profile_units, 0) if is_myvip else None
    )

    # Color rules per spec: red = "something is wrong", yellow = warning at 90%+
    def _capacity_color(actual: int, capacity: int) -> str:
        if capacity <= 0:
            return "0F172A"
        ratio = actual / capacity
        if ratio > 1.0:
            return RED_DEEP
        if ratio >= 0.9:
            return "B45309"  # amber
        return "0F172A"

    findings_rows = [
        (
            "Cars Registered on Property",
            cars_registered,
            _capacity_color(cars_registered, total_spaces),
        ),
    ]
    # Note: we don't currently distinguish reserved-space registrations from
    # general vehicle registrations in the parser. Show the metric only when we
    # have something meaningful to compare against.
    if reserved_spaces > 0:
        findings_rows.append((
            "Reserved Spaces Registered",
            "—",  # placeholder — needs reserved-spot tracking on the vehicle side
            "595959",
        ))
    if units_without_myvip is not None:
        findings_rows.append((
            "Units Without MyVIP Profile (TOW RISK)",
            units_without_myvip,
            RED_DEEP if units_without_myvip > 0 else "0F172A",
        ))

    af_start = af_band + 1
    for i, (label, value, color) in enumerate(findings_rows):
        r = af_start + i
        ws.row_dimensions[r].height = 24
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        lbl = ws.cell(row=r, column=2, value=label)
        lbl.font = Font(name=ARIAL, size=11, bold=False, color="404040")
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        val = ws.cell(row=r, column=6, value=value)
        val.font = Font(name=ARIAL, size=14, bold=True, color=color)
        val.alignment = Alignment(horizontal="right", vertical="center")

    findings_end = af_start + len(findings_rows) - 1

    # Vehicle Distribution sub-block (counts AND % of occupied units)
    vd_label_row = findings_end + 2
    ws.row_dimensions[vd_label_row].height = 16
    ws.merge_cells(start_row=vd_label_row, start_column=2, end_row=vd_label_row, end_column=7)
    vd_label = ws.cell(row=vd_label_row, column=2, value="Vehicle Distribution (per occupied unit)")
    vd_label.font = Font(name=ARIAL, size=10, bold=True, color=BRAND_BLUE)
    vd_label.alignment = Alignment(horizontal="left", vertical="center")

    distribution = audit.vehicle_distribution or {
        "0": 0, "1": 0, "2": 0, "3": 0, "4": 0, "5": 0, "6+": 0,
    }
    occupied_total = max(audit.occupied, 1)  # guard div/0
    bucket_keys = ["0", "1", "2", "3", "4", "5", "6+"]
    counts = [distribution.get(k, 0) for k in bucket_keys]

    # Header row of the distribution table
    vd_header_row = vd_label_row + 1
    ws.row_dimensions[vd_header_row].height = 18
    ws.cell(row=vd_header_row, column=2, value="Vehicles").font = Font(
        name=ARIAL, size=9, bold=True, color="595959",
    )
    for i, k in enumerate(bucket_keys):
        c = 3 + i
        h = ws.cell(row=vd_header_row, column=c, value=k)
        h.font = Font(name=ARIAL, size=10, bold=True, color="595959")
        h.alignment = Alignment(horizontal="center", vertical="center")

    # Counts row
    vd_count_row = vd_header_row + 1
    ws.row_dimensions[vd_count_row].height = 20
    ws.cell(row=vd_count_row, column=2, value="Units").font = Font(
        name=ARIAL, size=9, color="595959",
    )
    for i, n in enumerate(counts):
        c = 3 + i
        cell = ws.cell(row=vd_count_row, column=c, value=n)
        cell.font = Font(name=ARIAL, size=11, bold=True, color="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Percent row
    vd_pct_row = vd_count_row + 1
    ws.row_dimensions[vd_pct_row].height = 18
    ws.cell(row=vd_pct_row, column=2, value="% Occupied").font = Font(
        name=ARIAL, size=9, color="595959",
    )
    for i, n in enumerate(counts):
        c = 3 + i
        pct = (n / occupied_total) * 100 if audit.occupied else 0
        cell = ws.cell(row=vd_pct_row, column=c, value=f"{pct:.1f}%")
        cell.font = Font(name=ARIAL, size=10, color="595959")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── 5) Histogram (native Excel bar chart, brand-blue bars) ──────────
    chart_data_row = vd_pct_row + 2
    # Stash the bucket data in hidden cells so the chart can reference them.
    # Place them well below the cover content so they don't affect layout.
    data_anchor = chart_data_row + 30  # off the printed page, but still on sheet
    ws.cell(row=data_anchor, column=1, value="VehiclesPerUnit")
    ws.cell(row=data_anchor, column=2, value="Units")
    for i, k in enumerate(bucket_keys):
        ws.cell(row=data_anchor + 1 + i, column=1, value=k)
        ws.cell(row=data_anchor + 1 + i, column=2, value=counts[i])

    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = "Vehicles per Occupied Unit"
    chart.y_axis.title = "Number of Units"
    chart.x_axis.title = "Vehicles registered"
    chart.legend = None
    chart.height = 8
    chart.width = 16
    cats = Reference(
        ws, min_col=1, min_row=data_anchor + 1,
        max_row=data_anchor + len(bucket_keys),
    )
    vals = Reference(
        ws, min_col=2, min_row=data_anchor,
        max_row=data_anchor + len(bucket_keys),
    )
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    # Brand-blue bars + neutral-gray data labels
    for s in chart.series:
        s.graphicalProperties = GraphicalProperties(solidFill=BRAND_BLUE)
        s.graphicalProperties.line = LineProperties(solidFill=BRAND_BLUE_DARK)
    chart.dataLabels = DataLabelList(showVal=True)
    chart_anchor = f"B{chart_data_row}"
    ws.add_chart(chart, chart_anchor)

    # Print area covers the cover content (rows 2 → just past the chart).
    chart_end_row = chart_data_row + 18
    ws.print_area = f"A1:H{chart_end_row}"


# ------------------------------------------------------------------
# Sheet 2 — Full Cross-Reference
# ------------------------------------------------------------------
def _build_cross_reference(wb: Workbook, audit: AuditData) -> None:
    ws = wb.create_sheet("Cross-Reference")
    ws.sheet_properties.tabColor = MEDIUM_BLUE

    headers = [
        "Unit #", "Resident Name", "Occupancy Status",
        "In Parking System?", "# Vehicles Registered", "Audit Status",
    ]
    _apply_header_row(ws, headers, DARK_BLUE)

    status_styles = {
        STATUS_REGISTERED: (GREEN_FILL, Font(name=ARIAL, size=10, color=GREEN_TEXT)),
        STATUS_TOW_RISK: (RED_FILL, Font(name=ARIAL, size=10, bold=True, color=RED_TEXT)),
        STATUS_VACANT_HAS_PERMITS: (YELLOW_FILL, Font(name=ARIAL, size=10, color=YELLOW_TEXT)),
        STATUS_VACANT_OK: (YELLOW_FILL, Font(name=ARIAL, size=10, color=YELLOW_TEXT)),
        STATUS_PARKING_ONLY: (None, Font(name=ARIAL, size=10)),
    }

    for i, row in enumerate(audit.cross_ref):
        r = i + 2
        ws.cell(row=r, column=1, value=row.apt_display).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=row.name)
        ws.cell(row=r, column=3, value=row.occupancy)
        ws.cell(row=r, column=4, value="Yes" if row.in_parking_system else "No")
        ws.cell(row=r, column=5, value=row.vehicles).alignment = Alignment(horizontal="center")
        status_cell = ws.cell(row=r, column=6, value=row.audit_status)
        fill_color, font = status_styles.get(row.audit_status, (None, _data_font()))
        if fill_color:
            status_cell.fill = _fill(fill_color)
        status_cell.font = font

    end = 1 + len(audit.cross_ref)
    _set_data_borders(ws, 2, end, len(headers))
    _set_columns(ws, [10, 30, 18, 18, 22, 22])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(end, 2)}"


# ------------------------------------------------------------------
# Sheet 3 — Vehicle Details by Unit
# ------------------------------------------------------------------
def _build_vehicle_details(wb: Workbook, audit: AuditData) -> None:
    ws = wb.create_sheet("Vehicle Details")
    ws.sheet_properties.tabColor = "70AD47"

    # Per spec §3b — total vehicle count must be displayed at the top of this
    # sheet. Three header rows: total count + spacer + table header.
    total_row = 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    total_cell = ws.cell(
        row=total_row, column=1,
        value=f"Total Vehicles Registered: {audit.total_vehicles}",
    )
    total_cell.font = Font(name=ARIAL, size=14, bold=True, color="70AD47")
    total_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[total_row].height = 24

    header_row = 3
    headers = [
        "Unit #", "Resident (Rent Roll)", "Registered Name",
        "Make", "Model", "License Plate", "Date Registered",
    ]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.fill = _fill("70AD47")
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = DATA_BORDER

    # Build a quick lookup of rent-roll names by canonical apt
    rr_name = {row.apt_canonical: row.name for row in audit.cross_ref}

    r = header_row + 1
    for v in audit.vehicles:
        ws.cell(row=r, column=1, value=v.raw_apt or v.apt).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=rr_name.get(v.apt, ""))
        ws.cell(row=r, column=3, value=v.name)
        ws.cell(row=r, column=4, value=v.make)
        ws.cell(row=r, column=5, value=v.model)
        ws.cell(row=r, column=6, value=v.plate)
        ws.cell(row=r, column=7, value=v.date_registered)
        r += 1

    # NONE rows for occupied units with zero vehicles
    none_apts = getattr(audit, "_no_vehicle_apts", []) or []
    cross_by_apt = {row.apt_canonical: row for row in audit.cross_ref}
    none_font = Font(name=ARIAL, size=10, bold=True, color=RED_BRIGHT)
    for apt in none_apts:
        cr = cross_by_apt.get(apt)
        if not cr:
            continue
        ws.cell(row=r, column=1, value=cr.apt_display).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=cr.name)
        for col in (3, 4, 5, 6, 7):
            cell = ws.cell(row=r, column=col, value="NONE")
            cell.font = none_font
        r += 1

    end = r - 1
    _set_data_borders(ws, header_row + 1, end, len(headers))
    _set_columns(ws, [10, 28, 28, 12, 14, 16, 18])
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(end, header_row + 1)}"


# ------------------------------------------------------------------
# Sheet 4 — Vacant & Future Residents
# ------------------------------------------------------------------
def _build_vacant_future(wb: Workbook, audit: AuditData) -> None:
    ws = wb.create_sheet("Vacant & Future")
    ws.sheet_properties.tabColor = "FFC000"

    headers = ["Unit #", "Status", "Resident Name", "In Parking System?", "# Vehicles"]
    _apply_header_row(ws, headers, "FFC000")

    for i, v in enumerate(audit.vacants):
        r = i + 2
        ws.cell(row=r, column=1, value=v.apt_display).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=v.status_label)
        ws.cell(row=r, column=3, value=v.name)
        ws.cell(row=r, column=4, value="Yes" if v.in_parking_system else "No")
        ws.cell(row=r, column=5, value=v.vehicles).alignment = Alignment(horizontal="center")

    end = 1 + len(audit.vacants)
    _set_data_borders(ws, 2, end, len(headers))
    _set_columns(ws, [10, 18, 28, 20, 14])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(end, 2)}"


# ------------------------------------------------------------------
# Sheet 5 — Name Discrepancies
# ------------------------------------------------------------------
def _build_discrepancies(wb: Workbook, audit: AuditData) -> None:
    ws = wb.create_sheet("Name Discrepancies")
    ws.sheet_properties.tabColor = "FF6600"

    ws.cell(row=1, column=1, value=f"{audit.property_name} — NAME DISCREPANCIES").font = Font(
        name=ARIAL, size=14, bold=True,
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    subtitle = (
        f"Units where Rent Roll name does not match Parking Permit name | "
        f"{len(audit.discrepancies)} discrepancies found"
    )
    ws.cell(row=2, column=1, value=subtitle).font = Font(name=ARIAL, size=10, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

    headers = ["Unit", "Rent Roll Name", "Permit/Profile Name"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.fill = _fill(NAVY)
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center")
        cell.border = DATA_BORDER

    for i, d in enumerate(audit.discrepancies):
        r = 5 + i
        ws.cell(row=r, column=1, value=d.apt_display).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=d.rent_roll_name)
        permit_cell = ws.cell(row=r, column=3, value=d.permit_name)
        permit_cell.fill = _fill(ORANGE_HIGHLIGHT)
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.border = DATA_BORDER
            if not cell.font.name:
                cell.font = _data_font()

    # Total at bottom
    total_row = 5 + len(audit.discrepancies) + 1
    total_cell = ws.cell(
        row=total_row, column=1,
        value=f"Total discrepancies: {len(audit.discrepancies)}",
    )
    total_cell.font = Font(name=ARIAL, size=11, bold=True, color=RED_DEEP)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)

    _set_columns(ws, [10, 32, 32])
    ws.freeze_panes = "A5"
