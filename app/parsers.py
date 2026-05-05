"""Rent roll + vehicle data parsers (PDF, Excel, CSV).

Ported from the permit-audit skill — see SKILL.md "Data Parsing" section.

Each parser returns a normalized list of rows. The entry-point dispatch
chooses the right parser based on the filename extension.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Iterable, List, Optional, Tuple

from .audit_engine import AuditEngineError
from .normalize import extract_apt_number, is_skip_name


# ------------------------------------------------------------------
# Domain types
# ------------------------------------------------------------------
@dataclass
class RentRollEntry:
    apt: str                    # canonical (digits only, no leading zeros)
    raw_apt: str                # original from source
    name: str                   # raw name string
    status: str                 # 'occupied' | 'vacant' | 'model' | 'ntv'
    is_future: bool = False     # duplicate unit row = future/incoming resident


@dataclass
class Vehicle:
    apt: str                    # canonical
    raw_apt: str                # original
    name: str
    make: str = ""
    model: str = ""
    plate: str = ""
    date_registered: str = ""


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------
_NTV_RE = re.compile(r"\b(NTV|NOTICE\s*TO\s*VACATE|MOVE[-\s]*OUT)\b", re.I)
_VACANT_RE = re.compile(r"\bVACANT\b", re.I)
_MODEL_RE = re.compile(r"\bMODEL\b", re.I)


def _classify_status(name: str, status_text: str = "") -> str:
    blob = f"{name} {status_text}".upper()
    if _MODEL_RE.search(blob):
        return "model"
    if _NTV_RE.search(blob):
        return "ntv"
    if _VACANT_RE.search(blob) or not name.strip():
        return "vacant"
    return "occupied"


def _find_header_row(rows: Iterable[List[str]]) -> Tuple[int, dict]:
    """Locate header row (containing both Unit and Name columns) in tabular data.

    Returns (row_index, {col_role: col_index}). Roles: 'unit', 'name', 'status'.
    """
    rows_list = list(rows)
    for i, row in enumerate(rows_list[:30]):  # check first 30 rows max
        cells = [str(c or "").strip().lower() for c in row]
        unit_idx = name_idx = status_idx = None
        for j, c in enumerate(cells):
            if unit_idx is None and ("unit" in c or "apt" in c or "apartment" in c):
                unit_idx = j
            elif name_idx is None and ("name" in c or "resident" in c or "tenant" in c):
                name_idx = j
            elif status_idx is None and ("status" in c or "occupancy" in c):
                status_idx = j
        if unit_idx is not None and name_idx is not None:
            return i, {"unit": unit_idx, "name": name_idx, "status": status_idx}
    return -1, {}


# ------------------------------------------------------------------
# Rent roll — PDF
# ------------------------------------------------------------------
def parse_rent_roll_pdf(blob: bytes) -> List[RentRollEntry]:
    """Parse a rent-roll PDF using pdfplumber.

    Strategy:
      1. Try table extraction (some properties export structured tables).
      2. Fall back to **word-coordinate** extraction: group words by Y-axis
         to reconstruct rows, then walk each row left-to-right looking for
         a unit number followed by a resident name. Robust across formats
         where the name sits in the middle of a row separated by other
         tokens (typical Yardi/RealPage exports).
    """
    import pdfplumber

    entries: List[RentRollEntry] = []
    seen_units: set = set()

    try:
        with pdfplumber.open(io.BytesIO(blob)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                row_consumed = False
                for table in tables:
                    if not table:
                        continue
                    header_idx, cols = _find_header_row(table)
                    if header_idx < 0:
                        continue
                    row_consumed = True
                    for row in table[header_idx + 1:]:
                        if not row:
                            continue
                        raw_unit = (row[cols["unit"]] or "").strip() if cols.get("unit") is not None else ""
                        name = (row[cols["name"]] or "").strip() if cols.get("name") is not None else ""
                        status_text = (row[cols["status"]] or "").strip() if cols.get("status") is not None else ""
                        canonical = extract_apt_number(raw_unit)
                        if not canonical:
                            continue
                        is_future = canonical in seen_units
                        seen_units.add(canonical)
                        entries.append(RentRollEntry(
                            apt=canonical,
                            raw_apt=raw_unit,
                            name=name,
                            status=_classify_status(name, status_text),
                            is_future=is_future,
                        ))

                if not row_consumed:
                    entries.extend(_parse_rent_roll_words(page, seen_units))
    except AuditEngineError:
        raise
    except Exception as exc:
        raise AuditEngineError(
            code="parse_failed_rent_roll",
            message=f"Failed to read rent roll PDF: {exc}",
        )

    if not entries:
        raise AuditEngineError(
            code="parse_failed_rent_roll",
            message="Rent roll PDF parsed but no apartment rows were found.",
        )
    return entries


# Regex used to identify the leftmost token on a row that looks like an
# apartment unit (3-4 digits, with optional letter prefix like "A").
_UNIT_TOKEN_RE = re.compile(r"^[A-Za-z]?[0-9]{3,4}[A-Za-z]?$")
# Resident-id token (Yardi) — looks like t0656891. Used to anchor the name column.
_RESIDENT_ID_RE = re.compile(r"^t\d{5,}$", re.I)
# Tokens that are CLEARLY not name parts (dollar amounts, dates, percentages)
_NON_NAME_RE = re.compile(r"^[\d.,/$%-]+$")


def _row_words(page, y_tolerance: float = 3.0) -> List[List[dict]]:
    """Group words from a page into rows by approximate Y-coordinate."""
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False) or []
    if not words:
        return []
    words.sort(key=lambda w: (round(w["top"] / y_tolerance), w["x0"]))
    rows: List[List[dict]] = []
    current_y: Optional[float] = None
    current_row: List[dict] = []
    for w in words:
        if current_y is None or abs(w["top"] - current_y) > y_tolerance:
            if current_row:
                rows.append(current_row)
            current_row = [w]
            current_y = w["top"]
        else:
            current_row.append(w)
    if current_row:
        rows.append(current_row)
    return rows


def _parse_rent_roll_words(page, seen_units: set) -> List[RentRollEntry]:
    """Walk word-grouped rows and pull out (unit, name, status) triples.

    Looks for: leftmost token = unit number; then the resident name appears
    after the resident-id token (e.g. 't0656891') OR for vacant rows after
    'VACANT' tokens. We stop scanning when we hit a numeric token (rent/dep
    amounts).
    """
    out: List[RentRollEntry] = []
    for row in _row_words(page):
        if not row:
            continue
        tokens = [w["text"] for w in row]
        first = tokens[0].strip(",.")
        if not _UNIT_TOKEN_RE.match(first):
            continue
        raw_unit = first
        canonical = extract_apt_number(raw_unit)
        if not canonical:
            continue

        # Find anchor: resident-id token, or the first 'VACANT' / 'MODEL' marker
        anchor_idx = -1
        is_vacant_marker = False
        for i, tok in enumerate(tokens):
            up = tok.upper()
            if _RESIDENT_ID_RE.match(tok):
                anchor_idx = i
                break
            if up in {"VACANT", "MODEL"}:
                anchor_idx = i
                is_vacant_marker = True
                break

        # Collect name tokens AFTER the anchor (or after the unit type if no anchor)
        start_idx = anchor_idx + 1 if anchor_idx >= 0 else 2
        name_parts: List[str] = []
        for tok in tokens[start_idx:]:
            up = tok.upper()
            if up in {"VACANT", "MODEL"}:
                # second VACANT marker or status hint — keep capturing one more iteration to be safe
                if not name_parts:
                    name_parts.append(tok)
                break
            # Pure-numeric tokens (sqft, rent, deposit) — when we haven't
            # started collecting a name yet, this is almost certainly the
            # SqFt column wedged between unit-type and resident in Yardi
            # exports. Skip it and keep looking for the actual name. Once
            # we've started capturing, a numeric token marks end-of-name.
            if _NON_NAME_RE.match(tok):
                if not name_parts:
                    continue
                break
            # Mixed alphanumeric (plates, unit-type codes like '62842b2',
            # parenthetical amounts like 'Rodriguez(60),C'). Same rule —
            # skip while we have no name yet, otherwise treat as end-of-name.
            if any(c.isdigit() for c in tok):
                if not name_parts:
                    continue
                break
            name_parts.append(tok)

        name = " ".join(name_parts).strip()
        status_text = " ".join(tokens)
        is_future = canonical in seen_units
        seen_units.add(canonical)
        out.append(RentRollEntry(
            apt=canonical,
            raw_apt=raw_unit,
            name="" if (is_vacant_marker and not name_parts) else name,
            status=_classify_status(name, status_text),
            is_future=is_future,
        ))
    return out


# ------------------------------------------------------------------
# Rent roll — Excel
# ------------------------------------------------------------------
def _load_workbook_rows(blob: bytes, filename: str) -> List[List]:
    """Load worksheet rows from .xlsx (openpyxl) or legacy .xls (xlrd).

    Centralized so both rent-roll and vehicle workbook parsers handle the
    legacy binary BIFF format too. Many property managers still export
    .xls because their PMS pre-dates the 2007 transition. Without this,
    those uploads fail with a confusing "File is not a zip file" error.
    """
    name = (filename or "").lower()
    if name.endswith(".xls"):
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise AuditEngineError(
                code="parse_failed_rent_roll",
                message=("Legacy .xls support requires the xlrd package. "
                         "Re-save the file as .xlsx in Excel and re-upload."),
            ) from exc
        try:
            book = xlrd.open_workbook(file_contents=blob)
            sheet = book.sheet_by_index(0)
            rows: List[List] = []
            for r in range(sheet.nrows):
                rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
            return rows
        except AuditEngineError:
            raise
        except Exception as exc:
            raise AuditEngineError(
                code="parse_failed_rent_roll",
                message=f"Failed to open legacy .xls workbook: {exc}",
            )
    # Default: openpyxl path for .xlsx / .xlsm
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
        ws = wb.active
        return [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
    except Exception as exc:
        raise AuditEngineError(
            code="parse_failed_rent_roll",
            message=f"Failed to open rent roll workbook: {exc}",
        )


def parse_rent_roll_xlsx(blob: bytes, filename: str = "") -> List[RentRollEntry]:
    try:
        rows = _load_workbook_rows(blob, filename)
    except AuditEngineError:
        raise
    except Exception as exc:
        raise AuditEngineError(
            code="parse_failed_rent_roll",
            message=f"Failed to open rent roll workbook: {exc}",
        )

    header_idx, cols = _find_header_row(rows)
    if header_idx < 0:
        raise AuditEngineError(
            code="parse_failed_rent_roll",
            message="Could not find Unit/Name header row in rent roll.",
        )

    # Column-shift fallback: some Yardi exports (especially legacy .xls
    # round-trips) have merged-cell headers where the visible label sits
    # in a different column than the actual data. We sanity-check the
    # detected unit column against the next 20 data rows. If most rows
    # have something in unit_idx-1 but nothing in unit_idx, shift left.
    def _shift_if_misaligned(role: str, max_shift: int = 2) -> None:
        idx = cols.get(role)
        if idx is None or idx == 0:
            return
        sample = rows[header_idx + 1: header_idx + 21]
        def col_filled(rs, ci):
            return sum(1 for r in rs if ci < len(r) and str(r[ci]).strip())
        here = col_filled(sample, idx)
        for shift in range(1, max_shift + 1):
            new_idx = idx - shift
            if new_idx < 0:
                break
            there = col_filled(sample, new_idx)
            # Require the alternative column to be substantially better-filled
            if there >= max(3, here * 2 + 1):
                cols[role] = new_idx
                return
    _shift_if_misaligned("unit")
    _shift_if_misaligned("name")

    entries: List[RentRollEntry] = []
    seen: set = set()
    for row in rows[header_idx + 1:]:
        raw_unit = str(row[cols["unit"]]).strip() if cols.get("unit") is not None else ""
        name = str(row[cols["name"]]).strip() if cols.get("name") is not None else ""
        status_text = str(row[cols["status"]]).strip() if cols.get("status") is not None else ""
        canonical = extract_apt_number(raw_unit)
        if not canonical:
            continue
        is_future = canonical in seen
        seen.add(canonical)
        entries.append(RentRollEntry(
            apt=canonical,
            raw_apt=raw_unit,
            name=name,
            status=_classify_status(name, status_text),
            is_future=is_future,
        ))

    if not entries:
        raise AuditEngineError(
            code="parse_failed_rent_roll",
            message="Rent roll workbook had no apartment rows after header.",
        )
    return entries


# ------------------------------------------------------------------
# Vehicle data — CSV
# ------------------------------------------------------------------
_VEHICLE_FIELDS = {
    "apt": ["apartment #", "apt", "apt #", "unit", "unit #", "apartment"],
    "name": ["name", "names", "resident", "owner", "leaseholder", "leaseholders"],
    "make": ["make", "vehicle make"],
    "model": ["model", "vehicle model"],
    "plate": ["license plate", "plate", "tag"],
    "date": ["created", "date registered", "registered", "date"],
}


def _resolve_columns(headers: List[str]) -> dict:
    """Map our field names → actual column names in this CSV/XLSX."""
    lower = [h.strip().lower() for h in headers]
    found = {}
    for role, candidates in _VEHICLE_FIELDS.items():
        for cand in candidates:
            if cand in lower:
                found[role] = headers[lower.index(cand)]
                break
    return found


def parse_vehicle_csv(blob: bytes) -> List[Vehicle]:
    try:
        text = blob.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise AuditEngineError(
                code="parse_failed_vehicle_data",
                message="Vehicle CSV has no header row.",
            )
        cols = _resolve_columns(list(reader.fieldnames))
        if "apt" not in cols:
            raise AuditEngineError(
                code="parse_failed_vehicle_data",
                message="Could not find apartment column in vehicle CSV.",
            )
        out: List[Vehicle] = []
        for row in reader:
            raw_apt = (row.get(cols["apt"]) or "").strip()
            canonical = extract_apt_number(raw_apt)
            if not canonical:
                continue
            out.append(Vehicle(
                apt=canonical,
                raw_apt=raw_apt,
                name=(row.get(cols.get("name", "")) or "").strip(),
                make=(row.get(cols.get("make", "")) or "").strip(),
                model=(row.get(cols.get("model", "")) or "").strip(),
                plate=(row.get(cols.get("plate", "")) or "").strip(),
                date_registered=(row.get(cols.get("date", "")) or "").strip(),
            ))
        return out
    except AuditEngineError:
        raise
    except Exception as exc:
        raise AuditEngineError(
            code="parse_failed_vehicle_data",
            message=f"Failed to read vehicle CSV: {exc}",
        )


def parse_vehicle_xlsx(blob: bytes, filename: str = "") -> List[Vehicle]:
    try:
        rows = _load_workbook_rows(blob, filename)
    except AuditEngineError as exc:
        # Re-tag the error code from rent_roll → vehicle_data so the UI shows
        # the right context to the user, but keep the message text intact.
        raise AuditEngineError(
            code="parse_failed_vehicle_data",
            message=exc.message,
        )
    except Exception as exc:
        raise AuditEngineError(
            code="parse_failed_vehicle_data",
            message=f"Failed to open vehicle workbook: {exc}",
        )
    if not rows:
        raise AuditEngineError(
            code="parse_failed_vehicle_data",
            message="Vehicle workbook is empty.",
        )

    headers = [str(c or "").strip() for c in rows[0]]
    cols = _resolve_columns(headers)
    if "apt" not in cols:
        raise AuditEngineError(
            code="parse_failed_vehicle_data",
            message="Could not find apartment column in vehicle workbook.",
        )
    apt_idx = headers.index(cols["apt"])

    def col_idx(role: str) -> Optional[int]:
        name = cols.get(role)
        return headers.index(name) if name else None

    name_idx = col_idx("name")
    make_idx = col_idx("make")
    model_idx = col_idx("model")
    plate_idx = col_idx("plate")
    date_idx = col_idx("date")

    out: List[Vehicle] = []
    for row in rows[1:]:
        cells = list(row) + [None] * (len(headers) - len(row))  # pad short rows
        raw_apt = str(cells[apt_idx] or "").strip()
        canonical = extract_apt_number(raw_apt)
        if not canonical:
            continue
        out.append(Vehicle(
            apt=canonical,
            raw_apt=raw_apt,
            name=str(cells[name_idx] or "").strip() if name_idx is not None else "",
            make=str(cells[make_idx] or "").strip() if make_idx is not None else "",
            model=str(cells[model_idx] or "").strip() if model_idx is not None else "",
            plate=str(cells[plate_idx] or "").strip() if plate_idx is not None else "",
            date_registered=str(cells[date_idx] or "").strip() if date_idx is not None else "",
        ))
    return out


# ------------------------------------------------------------------
# Public dispatcher
# ------------------------------------------------------------------
def parse_rent_roll(blob: bytes, filename: str) -> List[RentRollEntry]:
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        return parse_rent_roll_pdf(blob)
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xls"):
        return parse_rent_roll_xlsx(blob, filename)
    raise AuditEngineError(
        code="validation_failed",
        message=f"Unsupported rent roll file type: {filename}",
    )


def parse_vehicle_data(blob: bytes, filename: str) -> List[Vehicle]:
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        return parse_vehicle_csv(blob)
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xls"):
        return parse_vehicle_xlsx(blob, filename)
    raise AuditEngineError(
        code="validation_failed",
        message=f"Unsupported vehicle data file type: {filename}",
    )
